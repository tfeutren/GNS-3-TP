# -------------------------- #
# ** Lire le ficher excel ** #
# -------------------------- #

import json
from openpyxl import load_workbook

# ---------------------------------------------
# ** FONCTIONS ** #

# Charger le fichier Excel
wb = load_workbook('Def_Root_AS.xlsx')
ws = wb.active

# Parcourez les lignes et les colonnes
def creerMatriceTableau():
    L = []
    M=[]
    for row in ws.iter_rows():
        for cell in row:
            L.append(cell.value)
        M.append(L)
        L = []
    return M

# Creer les strings pour les fichiers json

"Pour les Routeurs"
def creerStringJsonRoot(Tableur):
    lignesJSON = []
    for i in range(1, len(Tableur)):
        lignesJSON.append("{} = Routeur([{},{}],{},{},{},{})".format(Tableur[i][0],Tableur[i][1],Tableur[i][2],Tableur[i][3],Tableur[i][4],Tableur[i][5],Tableur[i][6]))
    return lignesJSON

"Pour les AS"
def creerStringJsonAS(Tableur):
    lignesJSON = []
    for i in range(1,len(Tableur)):
        if Tableur[i][8] != None:
            lignesJSON.append("{} = AS( {},{},{},{} ) ".format(Tableur[i][8],Tableur[i][9],Tableur[i][10],Tableur[i][11],Tableur[i][12]))
    return lignesJSON

# ---------------------------------------------
# ** PRINCIPAL ** #

Tableur = creerMatriceTableau()

List_AS = creerStringJsonAS(Tableur)
List_Router = creerStringJsonRoot(Tableur)



